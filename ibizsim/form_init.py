# -*-coding=utf-8-*-
from openpyxl import load_workbook
import os

# 设定提交表格的位置和名字
# 硬编码，需要修改
path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
filename = path+os.sep+'form'+os.sep+'2019_bizsim.xlsx'

# 本公司序数
company_number_row = 4
company_number_col = 4
# 总共公司数
total_company_raw = 4
total_company_col = 7
# 产品运出率
product_delivery_rate_row = 6
product_delivery_rate_col = 4
# 固定运费
fixed_freight_low_row = 9
fixed_freight_up_row = 11
fixed_freight_low_col = 4
fixed_freight_up_col = 12
# 变动运费
variable_freight_low_row = 13
variable_freight_up_row = 15
variable_freight_low_col = 4
variable_freight_up_col = 12
# 单位原材料库存费
per_material_inventory_cost_row = 16
per_material_inventory_cost_col = 4
# 成品库存费
product_inventory_cost_low_row = 19
product_inventory_cost_up_row = 23
product_inventory_cost_col = 4
# 单位生产资源
production_resources_low_row = 26
production_resources_up_row = 29
production_resources_low_col = 4
production_resources_up_col = 8
# 机器价格
machine_price_row = 30
machine_price_col = 4
# 折旧率
depreciation_rate_row = 30
depreciation_rate_col = 6
# 原材料价格
material_price_low_row = 34
material_price_up_row = 38
material_price_low_col = 4
material_price_up_col = 6
# 原材料运费
material_freight_low_row = 39
material_freight_up_row = 41
material_freight_col = 5
# 原材料使用率
material_usage_row = 41
material_usage_col = 4
# 管理费
management_fee_low_row = 44
management_fee_up_row = 48
management_fee_low_col = 4
management_fee_up_col = 6
# 机器维修费
maintenance_row = 49
maintenance_col = 4
# 研发费用
development_fee_low_row = 52
development_fee_up_row = 56
development_fee_low_col = 4
development_fee_up_col = 9
# 工人数据
workers_data_low_row = 57
workers_data_up_row = 63
workers_data_col = 4
# 工人工资
wages_row = 65
wages_low_col = 4
wages_up_col = 8
# 银行相关数据
bank_data_low_row = 67
bank_data_up_row = 73
bank_data_col = 4
# 偿债券率
bond_rate_row = 71
bond_rate_col = 6
# 评判标准
criteria_row = 75
criteria_low_col = 4
criteria_up_col = 11


# 写入比赛参数
def write_data(filename, data):
    try:
        excel = load_workbook(filename)
        sheetnames = excel.sheetnames
        table = excel[sheetnames[1]]
        n = 0

        # 本公司序数
        table.cell(company_number_row, company_number_col).value = data[n]
        n += 1
        # 总共公司数
        table.cell(total_company_raw, total_company_col).value = data[n]
        n += 1
        # 产品运出率
        table.cell(
            product_delivery_rate_row, product_delivery_rate_col).value = data[n]
        n += 1
        # 固定运费
        for row in range(fixed_freight_low_row, fixed_freight_up_row):
            for col in range(fixed_freight_low_col, fixed_freight_up_col):
                table.cell(row, col).value = data[n]
                n += 1
        # 变动运费
        for row in range(variable_freight_low_row, variable_freight_up_row):
            for col in range(variable_freight_low_col, variable_freight_up_col):
                table.cell(row, col).value = data[n]
                n += 1
        # 单位原材料库存费
        table.cell(
            per_material_inventory_cost_row, per_material_inventory_cost_col).value = data[n]
        n += 1
        # 成品库存费
        col = product_inventory_cost_col
        for row in range(product_inventory_cost_low_row, product_inventory_cost_up_row):
            table.cell(row, col).value = data[n]
            n += 1
        # 单位生产资源
        for row in range(production_resources_low_row, production_resources_up_row):
            for col in range(production_resources_low_col, production_resources_up_col):
                table.cell(row, col).value = data[n]
                n += 1
        # 机器价格
        table.cell(
            machine_price_row, machine_price_col).value = data[n]
        n += 1
        # 折旧率
        table.cell(
            depreciation_rate_row, depreciation_rate_col).value = data[n]
        n += 1
        # 原材料价格
        for row in range(material_price_low_row, material_price_up_row):
            for col in range(material_price_low_col, material_price_up_col):
                table.cell(row, col).value = data[n]
                n += 1
        # 原材料运费
        col = material_freight_col
        for row in range(material_freight_low_row, material_freight_up_row):
            table.cell(row, col).value = data[n]
            n += 1
        # 原材料使用率
        table.cell(
            material_usage_row, material_usage_col).value = data[n]
        n += 1
        # 管理费
        for row in range(management_fee_low_row, management_fee_up_row):
            for col in range(management_fee_low_col, management_fee_up_col):
                table.cell(row, col).value = data[n]
                n += 1
        # 机器维修费
        table.cell(maintenance_row, maintenance_col).value = data[n]
        n += 1
        # 研发费用
        for row in range(development_fee_low_row, development_fee_up_row):
            for col in range(development_fee_low_col, development_fee_up_col):
                table.cell(row, col).value = data[n]
                n += 1
        # 工人数据
        col = workers_data_col
        for row in range(workers_data_low_row, workers_data_up_row):
            table.cell(row, col).value = data[n]
            n += 1
        # 工人工资
        row = wages_row
        for col in range(wages_low_col, wages_up_col):
            table.cell(row, col).value = data[n]
            n += 1
        # 银行相关数据
        col = bank_data_col
        for row in range(bank_data_low_row, bank_data_up_row):
            table.cell(row, col).value = data[n]
            n += 1
        # 偿债券率
        table.cell(bond_rate_row, bond_rate_col).value = data[n]
        n += 1
        # 评判标准
        row = criteria_row
        for col in range(criteria_low_col, criteria_up_col):
            table.cell(row, col).value = data[n]
            n += 1
        # 保存
        excel.save(filename)
        print('sucessfully saved')
    except Exception, e:
        print str(e)


if __name__ == "__main__":
    import init_data
    team_name = u'全国熬夜锦标赛冠军选手'
    team_id = '350065'
    game_id = '177395'
    datas = init_data.get_init_data(team_name, game_id, team_id)
    write_data(filename, datas)
