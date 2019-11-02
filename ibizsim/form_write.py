# -*-coding=utf-8-*-
from openpyxl import load_workbook
from settings import *
import os

# 设定提交表格的位置和名字
# 硬编码，需要修改
path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
filename = path+os.sep+'form'+os.sep+'2019_bizsim.xlsx'

# 更新往期数据
def update_excel_data(filename, data, current_period):
    try:
        excel = load_workbook(filename)
        sheetnames = excel.sheetnames
        sheet_index = current_period-6
        table = excel[sheetnames[sheet_index]]
        n = 0

        # 上期发债卷数
        table.cell(bond_row, bond_col).value = data[n]
        n += 1

        # 上期债偿本金
        table.cell(principal_row, principal_col).value = data[n]
        n += 1

        # 上期期末企业状况
        for row in range(conditions_low_row, conditions_up_row):
            for col in range(conditions_low_col, conditions_up_col):
                table.cell(row, col).value = data[n]
                n += 1

        # 上期期末产品状况1
        for row in range(product1_low_row, product1_up_row):
            for col in range(product1_low_col, product1_up_col):
                table.cell(row, col).value = data[n]
                n += 1

        # 上期期末产品状况2
        for row in range(product2_low_row, product2_up_row):
            for col in range(product2_low_col, product2_up_col):
                table.cell(row, col).value = data[n]
                n += 1

        # 上期公共报表价格
        for row in range(price_low_row, price_up_row):
            for col in range(price_low_col, price_up_col):
                table.cell(row, col).value = data[n]
                n += 1

        # 上期市场份额
        for row in range(market_low_row, market_up_row):
            for col in range(market_low_col, market_up_col):
                table.cell(row, col).value = data[n]
                n += 1

        # 上期主要指标
        for row in range(indicators_low_row, indicators_up_row):
            for col in range(indicators_low_col, indicators_up_col):
                table.cell(row, col).value = data[n]
                n += 1
        # 保存
        excel.save(filename)
        print('sucessfully saved')
    except Exception, e:
        print str(e)

    

if __name__ == "__main__":

    import update_data
    team_name = u'全国熬夜锦标赛冠军选手'
    team_id = '350065'
    game_id = '177395'
    period_id = '3376324'
    company_number = '19'
    datas = update_data.get_update_data(team_name, game_id, team_id, period_id, company_number)
    update_excel_data(filename, datas, 15)
    print('success')