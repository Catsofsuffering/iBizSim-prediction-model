# -*-coding=utf-8-*-
from openpyxl import load_workbook
from settings import *
import os

# 设定提交表格的位置和名字
# 硬编码，需要修改
path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
filename = path+os.sep+'form'+os.sep+'2019_bizsim.xlsx'

def write_data(filename, current_period):
    try:
        # 打开Excel文件读取数据
        excel = load_workbook(filename)
        sheetnames = excel.sheetnames
        # 获取对应工作表
        sheet_index = current_period-6
        table = excel[sheetnames[sheet_index]]

        # 价格
        for row in range(price_low_row, price_up_row):
            for col in range(price_low_col, price_up_col):
                cell_value = table.cell(row, col).value
                price_list.append(cell_value)
        # 促销
        row = promotion_row
        for col in range(promotion_low_col, promotion_up_col):
            cell_value = table.cell(row, col).value
            promotion_list.append(cell_value)

        # 供货
        for row in range(support_low_row, support_up_row):
            for col in range(support_low_col, support_up_col):
                cell_value = table.cell(row, col).value
                support_list.append(cell_value)

        # 生产
        for row in range(product_low_row, product_up_row):
            for col in range(product_low_col, product_up_col):
                cell_value = table.cell(row, col).value
                product_list.append(cell_value)

        # 发展
        row = develop_row
        for col in range(develop_low_col, develop_up_col):
            cell_value = table.cell(row, col).value
            develop_list.append(cell_value)

        # 财务
        row = finance_row
        for col in range(finance_low_col, finance_up_col):
            cell_value = table.cell(row, col).value
            finance_list.append(cell_value)

        # test
        # print(len(price_list))
        # print(len(promotion_list))
        # print(len(support_list))
        # print(len(product_list))
        # print(len(develop_list))
        # print(len(finance_list))

        # 合并返回
        result_list = price_list + support_list + promotion_list + \
            product_list + develop_list + finance_list
        return result_list

    except Exception, e:
        print str(e)


if __name__ == "__main__":
    data = write_data(filename, 9)
    # print(data)
    for i in data:
        print(i)
